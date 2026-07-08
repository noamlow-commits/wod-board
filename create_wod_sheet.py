import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "WOD"
ws.sheet_view.rightToLeft = True

# Column widths
ws.column_dimensions['A'].width = 15
ws.column_dimensions['B'].width = 40
ws.column_dimensions['C'].width = 45
ws.column_dimensions['D'].width = 35
ws.column_dimensions['E'].width = 30

# Colors
green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
light_green_fill = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")
yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
orange_fill = PatternFill(start_color="F0C27A", end_color="F0C27A", fill_type="solid")
header_font = Font(bold=True, size=14)
safety_font = Font(bold=True, size=16, color="FF0000")
label_font = Font(bold=True, size=14)
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Row 1 - Headers
headers = ["", "1", "2", "3", "בטיחות/דגשים"]
for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.fill = green_fill
    cell.border = thin_border

# Row 2 - WOD
ws.cell(row=2, column=1, value="WOD")
ws.cell(row=2, column=1).font = label_font
ws.cell(row=2, column=1).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
ws.cell(row=2, column=1).border = thin_border

# Column B (1)
ws.cell(row=2, column=2, value="2 rounds\n400 m run\n10 kang squat> with barbell\n30 seconds bent over stretch on wall")
ws.cell(row=2, column=2).alignment = Alignment(wrap_text=True, vertical='top')
ws.cell(row=2, column=2).fill = green_fill
ws.cell(row=2, column=2).border = thin_border

# Column C (2)
ws.cell(row=2, column=3, value="10 min power clean tech and complex warm up\n\nE2MOM x 6\n1 high pull\n1 power clean\n2 front squat")
ws.cell(row=2, column=3).alignment = Alignment(wrap_text=True, vertical='top')
ws.cell(row=2, column=3).fill = green_fill
ws.cell(row=2, column=3).border = thin_border

# Column D (3)
ws.cell(row=2, column=4, value="18 min amrap\n8.12b\n10 db box step up\n12 db/bar clean  22.5/15 (1 db)")
ws.cell(row=2, column=4).alignment = Alignment(wrap_text=True, vertical='top')
ws.cell(row=2, column=4).fill = green_fill
ws.cell(row=2, column=4).border = thin_border

# Column E (בטיחות/דגשים)
ws.cell(row=2, column=5, value="לא זורקים מוטות ריקים/\nעם חמישיות")
ws.cell(row=2, column=5).font = safety_font
ws.cell(row=2, column=5).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
ws.cell(row=2, column=5).fill = yellow_fill
ws.cell(row=2, column=5).border = thin_border

# Row height for WOD
ws.row_dimensions[2].height = 130

# Row 3 - CARDIO
ws.cell(row=3, column=1, value="CARDIO")
ws.cell(row=3, column=1).font = label_font
ws.cell(row=3, column=1).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
ws.cell(row=3, column=1).border = thin_border

# CARDIO content in column C (2)
ws.cell(row=3, column=2, value="")
ws.cell(row=3, column=2).fill = orange_fill
ws.cell(row=3, column=2).border = thin_border

ws.cell(row=3, column=3, value="E2MOM X 20\nStay 5 sets in the same station\n\nA - 20/16 Cal ROW.\nB - 15-25 Burpees.\nC - 15 Box Jumps + 15 Sit Ups\nD- 300 M RUN")
ws.cell(row=3, column=3).alignment = Alignment(wrap_text=True, vertical='top')
ws.cell(row=3, column=3).fill = orange_fill
ws.cell(row=3, column=3).border = thin_border

ws.cell(row=3, column=4, value="")
ws.cell(row=3, column=4).fill = orange_fill
ws.cell(row=3, column=4).border = thin_border

ws.cell(row=3, column=5, value="")
ws.cell(row=3, column=5).fill = orange_fill
ws.cell(row=3, column=5).border = thin_border

# Row height for CARDIO
ws.row_dimensions[3].height = 150

output_path = r"c:\Users\User\Desktop\WOD_sheet.xlsx"
wb.save(output_path)
print(f"Saved to {output_path}")
